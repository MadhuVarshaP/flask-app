
from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
import cv2
from ultralytics import YOLO
import openpyxl
from datetime import datetime
import os

app = Flask(__name__)
CORS(app)

model_path = 'models/Freshnew100.pt'  # Path to your .pt model file
model = YOLO(model_path)  # Load the YOLO model


# Define label mapping
label_encoder = ['apple_fresh', 'apple_stale', 'onion_fresh', 'onion_stale', 
                 'carrot_fresh', 'carrot_stale', 'tomato_fresh', 'tomato_stale']

# Define expected lifespan for each product
expected_life_span = {
    "apple": 7, "onion": 10, "carrot": 5, "tomato": 3
}

# Confidence threshold
CONFIDENCE_THRESHOLD = 0.5

# Excel file for storing fresh count
excel_file = "detection_fresh_count3.xlsx"

# Initialize or load Excel workbook and sheet
try:
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["S No", "Product", "Fresh Count", "Last Detected Time", "Expected Life Span"])

# Helper function to update fresh count in Excel
def update_fresh_count(product, is_fresh):
    lifespan = "N/A" if not is_fresh else expected_life_span.get(product, "Unknown")
    product_found = False
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[1].value == product:
            product_found = True
            if is_fresh:
                row[2].value += 1
            row[3].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            row[4].value = lifespan
            break
    if not product_found:
        fresh_count = 1 if is_fresh else 0
        last_detected_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([sheet.max_row, product, fresh_count, last_detected_time, lifespan])

# Route for rendering the index page
@app.route('/')
def home():
    return render_template('index.html')


# Route for handling freshness detection
@app.route('/detect-freshness', methods=['POST'])
def detect_freshness():
    try:
        if 'image' not in request.files:
            return jsonify({"error": "No image file provided"}), 400

        image_file = request.files['image']
        uploads_dir = "uploads"
        os.makedirs(uploads_dir, exist_ok=True)  # Ensure the directory exists
        image_path = os.path.join(uploads_dir, image_file.filename)

        # Save the image temporarily
        image_file.save(image_path)

        # Process the image using YOLO
        image = cv2.imread(image_path)
        if image is None:
            return jsonify({"error": "Error reading image"}), 400

        results = model(image)
        detection_results = []

        for result in results:
            boxes = result.boxes.xyxy  # Bounding box coordinates
            confidences = result.boxes.conf  # Confidence scores
            labels = result.boxes.cls  # Class indices

            for i, box in enumerate(boxes):
                confidence = confidences[i].item()
                if confidence < CONFIDENCE_THRESHOLD:
                    continue  # Skip low-confidence predictions

                x1, y1, x2, y2 = map(int, box)
                label_idx = int(labels[i])
                predicted_label = label_encoder[label_idx]
                product, freshness = predicted_label.split('_')

                # Update fresh count in Excel
                is_fresh = (freshness == "fresh")
                update_fresh_count(product, is_fresh)

                # Append result for the frontend
                detection_results.append({
                    "product": product,
                    "freshness": freshness,
                    "confidence": confidence,
                    "bbox": [x1, y1, x2, y2]
                })

        # Save the workbook
        workbook.save(excel_file)

        return jsonify({
            "status": "Freshness detection completed",
            "detections": detection_results
        })

    except Exception as e:
        print(f"Error in detect_freshness route: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/download-excel', methods=['GET'])
def download_excel():
    try:
        # Ensure the Excel file exists
        if not os.path.exists(excel_file):
            return jsonify({"error": "Excel file not found"}), 404

        # Return the Excel file as a downloadable response
        return send_file(excel_file, as_attachment=True, download_name="detection_fresh_count.xlsx")
    except Exception as e:
        print(f"Error in download_excel route: {e}")
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True,host='0.0.0.0',port=80)
